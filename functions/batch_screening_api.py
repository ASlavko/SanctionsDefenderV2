import functions_framework
import json
from google.cloud import firestore
from datetime import datetime
from flask import jsonify
import re
import io
import csv
from openpyxl import load_workbook
from matching import match_name

db = firestore.Client()

@functions_framework.http
def batch_screening(request):
    """
    Batch screening endpoint for uploaded Excel/CSV files.
    
    Accepts:
    - POST with multipart/form-data
    - File upload (Excel or CSV)
    - entity_type: 'company' (default) or 'individual'
    - threshold: fuzzy match threshold (0-100, default 80)
    
    Returns:
    - results: Array of screening results with matches
    - summary: Overall statistics
    - query_time_ms: Processing time
    """
    
    start_time = datetime.now()
    
    # Enable CORS
    if request.method == 'OPTIONS':
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'POST, OPTIONS',
            'Access-Control-Allow-Headers': 'Content-Type'
        }
        return ('', 204, headers)
    
    if request.method != 'POST':
        return jsonify({
            'error': 'Method not allowed. Use POST.',
            'status': 'invalid_request'
        }), 405
    
    try:
        # Get form parameters
        entity_type = request.form.get('entity_type', 'company').lower()
        threshold = int(request.form.get('threshold', 80))
        
        # Validate entity type
        if entity_type not in ['company', 'individual']:
            return jsonify({
                'error': 'entity_type must be "company" or "individual"',
                'status': 'invalid_request'
            }), 400
        
        # Get uploaded file
        if 'file' not in request.files:
            return jsonify({
                'error': 'No file uploaded. Please upload an Excel or CSV file.',
                'status': 'invalid_request'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'error': 'No file selected',
                'status': 'invalid_request'
            }), 400
        
        # Parse file and extract names
        names = _parse_file(file)
        
        if not names:
            return jsonify({
                'error': 'No names found in file. Please check file format.',
                'status': 'invalid_request'
            }), 400
        
        # Screen each name against sanctions database
        results = []
        for idx, name in enumerate(names):
            try:
                matches = _screen_name(name, entity_type, threshold)
                results.append({
                    'name': name,
                    'status': 'POTENTIAL_MATCH' if matches else 'CLEAR',
                    'match_count': len(matches),
                    'matches': matches
                })
            except Exception as e:
                # Log error but continue with other names
                print(f'[WARN] Error screening name "{name}": {str(e)}')
                results.append({
                    'name': name,
                    'status': 'ERROR',
                    'match_count': 0,
                    'matches': [],
                    'error': str(e)
                })
        
        # Calculate summary statistics
        summary = {
            'total_screened': len(results),
            'clear': sum(1 for r in results if r['status'] == 'CLEAR'),
            'potential_matches': sum(1 for r in results if r['status'] == 'POTENTIAL_MATCH'),
            'entity_type': entity_type,
            'threshold': threshold
        }
        
        # Calculate execution time
        exec_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        
        response = {
            'status': 'success',
            'results': results,
            'summary': summary,
            'query_time_ms': round(exec_time_ms, 2)
        }
        
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/json'
        }
        
        return (json.dumps(response, ensure_ascii=False), 200, headers)
    
    except Exception as e:
        exec_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return jsonify({
            'status': 'error',
            'error': str(e),
            'query_time_ms': round(exec_time_ms, 2)
        }), 500


def _parse_file(file):
    """Parse Excel or CSV file and extract names"""
    names = []
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.csv'):
            # Parse CSV
            content = file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                if row and row[0].strip():
                    # Take first column
                    names.append(row[0].strip())
        
        elif filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    # Take first column
                    name = str(row[0]).strip()
                    if name:
                        names.append(name)
        
        else:
            raise ValueError('Unsupported file format. Please upload CSV or Excel (.xlsx, .xls)')
    
    except Exception as e:
        raise ValueError(f'Error parsing file: {str(e)}')
    
    # Remove header row if it looks like a header
    if names and any(keyword in names[0].lower() for keyword in ['name', 'company', 'individual', 'entity']):
        names = names[1:]
    
    # Remove duplicates while preserving order
    seen = set()
    unique_names = []
    for name in names:
        name_lower = name.lower()
        if name_lower not in seen:
            seen.add(name_lower)
            unique_names.append(name)
    
    return unique_names


def _screen_name(name, entity_type, threshold):
    """Screen a single name against sanctions database"""
    matches = []
    
    # Normalize name for matching
    name_normalized = name.lower().strip()
    
    # Query Firestore with limit to prevent timeouts
    # Limit to 5000 documents per entity type for performance
    query = db.collection('sanctions_entities')
    query = query.where('entity_type', '==', entity_type)
    query = query.limit(5000)
    
    docs = query.stream()
    
    # Check each sanctions record
    for doc in docs:
        data = doc.to_dict()
        score = _calculate_fuzzy_score(name_normalized, data, entity_type)
        
        if score >= threshold:
            matches.append({
                'id': data.get('id'),
                'source': data.get('sanction_source'),
                'name': data.get('main_name'),
                'aliases': data.get('aliases', []),
                'entity_type': data.get('entity_type'),
                'country': data.get('country'),
                'programs': data.get('details', {}).get('programs', []),
                'confidence': round(score, 2)
            })
    
    # Sort by confidence (highest first)
    matches.sort(key=lambda x: x['confidence'], reverse=True)
    
    # Return top 5 matches
    return matches[:5]


def _calculate_fuzzy_score(query_name, record, entity_type):
    """
    Calculate fuzzy match score (0-100) using advanced matching algorithm.
    
    Uses multiple techniques:
    - Exact matching
    - Levenshtein distance (edit distance)
    - Phonetic matching (Soundex/Metaphone)
    - Token-based matching
    """
    return match_name(query_name, record, entity_type=entity_type, use_phonetic=True)

